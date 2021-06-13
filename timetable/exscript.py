from openpyxl import load_workbook
import os
import json

def open_spreadsheet(name):
    if not name:
        print("Err: open_spreadsheet: invalid file name")
        return 
    wb = load_workbook(filename=name)
    return wb

def mapvals(ls):
    return list(map(lambda cell:cell.value, ls))

def create_tt(sheet):
    print("create_tt: Working on :", sheet.title)
    num_rows = sheet.max_row
    num_cols = sheet.max_column
    print("Sheet size: ", num_rows, " X ", num_cols)
    columns = list(sheet.columns)
    rows = list(sheet.rows)
    days = columns[0][1:]
    print(mapvals(days))
    times = rows[0][1:]
    print(mapvals(times))
    tt = {}
    row_num = 1
    col_num = 1
    lmsr = list(sheet.merged_cells.ranges)
    for row in rows[1:]:
        print("Row :", row_num)
        col_num = 1
        dayname = days[row_num-1].value
        innertt = {}
        for time in times:
            innertt[time.value] = None
        while col_num < len(row):
            print("Column :", col_num)
            cell = row[col_num]
            print(cell.value)
            range_start = False
            range_info = None
            timeval = times[col_num-1].value
            # check if the cell is part of a merged range for every merged cell range
            for msr in lmsr:
                strt_cell = msr.start_cell
                if strt_cell == cell:
                    range_start = True
                    range_info = msr
                    break
            if range_start:
                sz = msr.size['columns']
                print('Range of size: ', sz, ' starting')
                if cell.value is not None:
                    for i in range(sz):
                        innertt[times[col_num - 1 + i].value] = [cell.value, sz - i]
                col_num += msr.size['columns'] 
                range_start = False
            else:
                if cell.value is not None:
                    innertt[timeval] = [cell.value, 1]
                col_num += 1
        tt[dayname] = innertt
        row_num += 1
    return tt

        
def ttmaker(year, batchgr, batchnum, toJSON=False):
    batchgr = batchgr.strip()
    year = year.strip()
    pwd = os.path.dirname(os.path.realpath(__file__))
    print('Current dir: ', pwd)
    filename = str(pwd) + "/ttdata/" + year.upper() + "-" + batchgr.upper() + ".xlsx"
    wb = open_spreadsheet(filename)
    tt = create_tt(wb[wb.sheetnames[3*(batchnum-1)]])
    if toJSON:
        return json.dumps(tt)
    return tt


if __name__=='__main__':
    print("Script Started...")
    # just for sheet 1 now
    tt = ttmaker("2k19", "A", 2)
    print(tt)

